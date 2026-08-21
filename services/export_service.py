"""
Unified export service — CSV, Excel, and PDF generation.

Replaces the previous 5 redundant export implementations
(``exports/csv_exporter.py``, ``exports/excel_exporter.py``,
``exports/excel_generator.py``, ``exports/pdf_exporter.py``,
``exports/pdf_generator.py``) with a single, typed, tested service.

Supports both **file‑system** and **in‑memory (bytes)** export.

Usage
-----
.. code:: python

    svc = ExportService()

    # Save to file
    path = svc.to_csv("students.csv", ["Name", "Grade"], data)
    path = svc.to_excel("students.xlsx", ["Name", "Grade"], data)
    path = svc.to_pdf("students.pdf", "Student Report", headers, data)

    # Bytes (for API responses)
    pdf_bytes = svc.to_pdf_bytes("Student Report", headers, data)
"""

import csv
import io
import os
from dataclasses import dataclass
from typing import Any, List, Optional, Sequence, Tuple, Union

import openpyxl
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from config.settings import BASE_DIR
from utils.logger import setup_logger

logger = setup_logger("export")

# Export directory

EXPORT_DIR = os.path.join(BASE_DIR, "exports", "generated")

# Types

Row = Sequence[Any]
HeaderRow = Sequence[str]


@dataclass
class ExportResult:
    """Result of an export operation.

    Attributes
    ----------
    path : str or None
        Filesystem path if saved to disk; ``None`` if bytes-only.
    bytes_ : bytes or None
        Raw file bytes if in-memory; ``None`` if saved to disk.
    filename : str
        The original filename requested.
    mime_type : str
        MIME content type for the exported file.
    """

    path: Optional[str]
    bytes_: Optional[bytes]
    filename: str
    mime_type: str


# Exceptions


class ExportError(Exception):
    """Raised when an export operation fails."""


# Service


class ExportService:
    """Generate CSV, Excel, and PDF exports.

    Parameters
    ----------
    export_dir : str, optional
        Directory to write exported files. Defaults to
        ``<project_root>/exports/generated/``.
    auto_create : bool
        Create *export_dir* on instantiation if it does not exist
        (default ``True``).
    """

    MIME_CSV = "text/csv"
    MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    MIME_PDF = "application/pdf"

    def __init__(self, export_dir: Optional[str] = None, auto_create: bool = True) -> None:
        self.export_dir = export_dir or EXPORT_DIR
        if auto_create:
            os.makedirs(self.export_dir, exist_ok=True)

    # ── Public API ────────────────────────────────────────────────

    def to_csv(
        self,
        filename: str,
        headers: HeaderRow,
        rows: List[Row],
    ) -> ExportResult:
        """Export data as CSV.

        Parameters
        ----------
        filename : str
            Output filename (e.g. ``"students.csv"``).
        headers : sequence of str
            Column header names.
        rows : list of sequences
            Data rows.

        Returns
        -------
        ExportResult
        """
        logger.info("Exporting CSV: %s (%d rows)", filename, len(rows))
        path = os.path.join(self.export_dir, filename)

        try:
            with open(path, mode="w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(headers)
                writer.writerows(rows)
        except OSError as exc:
            logger.error("CSV export failed: %s", exc)
            raise ExportError(f"Could not write CSV to {path}: {exc}") from exc

        logger.info("CSV exported: %s", path)
        return ExportResult(
            path=path,
            bytes_=None,
            filename=filename,
            mime_type=self.MIME_CSV,
        )

    def to_csv_bytes(
        self,
        headers: HeaderRow,
        rows: List[Row],
    ) -> ExportResult:
        """Export data as CSV in memory.

        Returns
        -------
        ExportResult with ``bytes_`` set.
        """
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(headers)
        writer.writerows(rows)
        raw = buf.getvalue().encode("utf-8")

        return ExportResult(
            path=None,
            bytes_=raw,
            filename="export.csv",
            mime_type=self.MIME_CSV,
        )

    # ── Excel ─────────────────────────────────────────────────────

    def to_excel(
        self,
        filename: str,
        headers: HeaderRow,
        rows: List[Row],
        sheet_name: str = "Sheet1",
    ) -> ExportResult:
        """Export data as Excel (``.xlsx``).

        Applies basic formatting: bold headers with a light blue fill.

        Parameters
        ----------
        filename : str
            Output filename (e.g. ``"students.xlsx"``).
        headers : sequence of str
            Column header names.
        rows : list of sequences
            Data rows.
        sheet_name : str
            Name of the worksheet (default ``"Sheet1"``).

        Returns
        -------
        ExportResult
        """
        logger.info("Exporting Excel: %s (%d rows)", filename, len(rows))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        self._style_excel_headers(ws, headers)
        for row in rows:
            ws.append(list(row))

        # Auto-adjust column widths
        self._auto_width(ws)

        path = os.path.join(self.export_dir, filename)
        try:
            wb.save(path)
        except OSError as exc:
            logger.error("Excel export failed: %s", exc)
            raise ExportError(f"Could not write Excel to {path}: {exc}") from exc

        logger.info("Excel exported: %s", path)
        return ExportResult(
            path=path,
            bytes_=None,
            filename=filename,
            mime_type=self.MIME_XLSX,
        )

    def to_excel_bytes(
        self,
        headers: HeaderRow,
        rows: List[Row],
        sheet_name: str = "Sheet1",
    ) -> ExportResult:
        """Export data as Excel in memory.

        Returns
        -------
        ExportResult with ``bytes_`` set.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        self._style_excel_headers(ws, headers)
        for row in rows:
            ws.append(list(row))
        self._auto_width(ws)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        return ExportResult(
            path=None,
            bytes_=buf.read(),
            filename="export.xlsx",
            mime_type=self.MIME_XLSX,
        )

    # ── PDF ───────────────────────────────────────────────────────

    def to_pdf(
        self,
        filename: str,
        title: str,
        headers: HeaderRow,
        rows: List[Row],
        page_size: Tuple[float, float] = A4,
        landscape: bool = False,
    ) -> ExportResult:
        """Export data as PDF with a styled table.

        Parameters
        ----------
        filename : str
            Output filename (e.g. ``"report.pdf"``).
        title : str
            Document title (rendered as a heading).
        headers : sequence of str
            Column header names.
        rows : list of sequences
            Data rows.
        page_size : tuple
            Page dimensions (default A4).
        landscape : bool
            If ``True``, swap width/height for landscape orientation.

        Returns
        -------
        ExportResult
        """
        logger.info("Exporting PDF: %s (%d rows)", filename, len(rows))

        if landscape:
            page_size = (page_size[1], page_size[0])

        path = os.path.join(self.export_dir, filename)

        try:
            self._build_pdf(path, title, headers, rows, page_size)
        except Exception as exc:
            logger.error("PDF export failed: %s", exc)
            raise ExportError(f"Could not write PDF to {path}: {exc}") from exc

        logger.info("PDF exported: %s", path)
        return ExportResult(
            path=path,
            bytes_=None,
            filename=filename,
            mime_type=self.MIME_PDF,
        )

    def to_pdf_bytes(
        self,
        title: str,
        headers: HeaderRow,
        rows: List[Row],
        page_size: Tuple[float, float] = A4,
        landscape: bool = False,
    ) -> ExportResult:
        """Export data as PDF in memory.

        Returns
        -------
        ExportResult with ``bytes_`` set.
        """
        if landscape:
            page_size = (page_size[1], page_size[0])

        buf = io.BytesIO()
        self._build_pdf(buf, title, headers, rows, page_size)

        return ExportResult(
            path=None,
            bytes_=buf.getvalue(),
            filename="export.pdf",
            mime_type=self.MIME_PDF,
        )

    # ── Convenience: detect format from filename ──────────────────

    def export(
        self,
        filename: str,
        headers: HeaderRow,
        rows: List[Row],
        **kwargs: Any,
    ) -> ExportResult:
        """Auto-detect format from *filename* extension and export.

        Supported extensions: ``.csv``, ``.xlsx``, ``.pdf``.

        Parameters
        ----------
        filename : str
            Output filename (e.g. ``"report.csv"``).
        headers : sequence of str
        rows : list of sequences
        **kwargs
            Passed through to the format-specific method.

        Returns
        -------
        ExportResult

        Raises
        ------
        ExportError
            If the extension is unsupported.
        """
        ext = os.path.splitext(filename)[1].lower()
        if ext == ".csv":
            return self.to_csv(filename, headers, rows, **kwargs)
        elif ext == ".xlsx":
            return self.to_excel(filename, headers, rows, **kwargs)
        elif ext == ".pdf":
            return self.to_pdf(filename, title="Export", headers=headers, rows=rows, **kwargs)
        else:
            raise ExportError(f"Unsupported export format '{ext}'. Supported: .csv, .xlsx, .pdf")

    # ── Internal helpers ──────────────────────────────────────────

    @staticmethod
    def _style_excel_headers(
        ws: openpyxl.worksheet.worksheet.Worksheet, headers: HeaderRow
    ) -> None:
        """Apply bold + light-blue fill to the header row."""
        header_fill = PatternFill(start_color="B4D6F7", end_color="B4D6F7", fill_type="solid")
        header_font = Font(bold=True, size=11)

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font

    @staticmethod
    def _auto_width(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """Approximate auto-width based on cell content length."""
        for col_cells in ws.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                try:
                    cell_len = len(str(cell.value or ""))
                    if cell_len > max_len:
                        max_len = cell_len
                except Exception:
                    pass
            adjusted = min(max_len + 3, 60)
            ws.column_dimensions[col_letter].width = adjusted

    def _build_pdf(
        self,
        destination: Union[str, io.BytesIO],
        title: str,
        headers: HeaderRow,
        rows: List[Row],
        page_size: Tuple[float, float],
    ) -> None:
        """Build a PDF document with title and data table.

        Parameters
        ----------
        destination : str or BytesIO
            File path or byte buffer to write to.
        title : str
        headers : sequence of str
        rows : list of sequences
        page_size : (width, height)
        """
        styles = getSampleStyleSheet()
        title_style = styles["Title"]

        doc = SimpleDocTemplate(
            destination,
            pagesize=page_size,
            topMargin=20 * mm,
            bottomMargin=20 * mm,
            leftMargin=15 * mm,
            rightMargin=15 * mm,
        )

        elements: List[Any] = []

        # Title
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 6 * mm))

        # Build table data
        table_data: List[List[Any]] = [list(headers)]
        for row in rows:
            table_data.append([str(cell) if cell is not None else "" for cell in row])

        # Column widths: distribute evenly
        usable_width = page_size[0] - 30 * mm  # minus margins
        n_cols = len(headers)
        col_width = usable_width / n_cols if n_cols else usable_width

        table = Table(table_data, colWidths=[col_width] * n_cols, repeatRows=1)

        # Style
        header_color = colors.HexColor("#2563EB")
        stripe_color = colors.HexColor("#F1F5F9")

        style_commands = [
            ("BACKGROUND", (0, 0), (-1, 0), header_color),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]

        # Alternating row colours
        for i in range(1, len(table_data)):
            if i % 2 == 0:
                style_commands.append(("BACKGROUND", (0, i), (-1, i), stripe_color))

        table.setStyle(TableStyle(style_commands))

        elements.append(table)
        doc.build(elements)
